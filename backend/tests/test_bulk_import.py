"""통합본(법인=시트) 파서와 일괄 판정 엔드포인트.

실제 통합본은 사람이 손으로 만든 파일이라 시트마다 표가 시작하는 열이 다르고,
아직 안 채워진 법인은 숫자 대신 '반기매출입력' 같은 안내 문구가 들어 있다.
여기서 검증하는 것은 그 지저분함을 **조용히 넘기지 않는지**다 —
못 읽은 법인을 0 으로 계산해 '해당없음'을 내는 것이 최악이다.
"""
import io

import pytest
from fastapi.testclient import TestClient

from backend import bulk_import
from backend.auth import ADMIN_USERNAME, ADMIN_PASSWORD
from backend.main import app

client = TestClient(app)

SUBJECT = "가나전자"
COUNTERPARTY = "자차산업"
OTHER = "마바물산"


def _sheet(ws, *, name, size, total, income, rows, start_col=4, tax=None):
    """실제 통합본과 같은 모양으로 한 시트를 그린다.

    `start_col` 로 표 시작 열을 옮길 수 있다 — 원본 파일이 시트마다 다르기 때문에
    열을 고정해 읽으면 절반이 빈 값으로 읽힌다.
    """
    c = start_col
    ws.cell(row=2, column=c, value="법인명")
    ws.cell(row=2, column=c + 3, value=name)
    ws.cell(row=4, column=c, value="기업구분")
    ws.cell(row=4, column=c + 3, value=size)
    ws.cell(row=6, column=c, value="총매출액")
    ws.cell(row=6, column=c + 3, value=total)
    ws.cell(row=8, column=c, value="영업이익")
    ws.cell(row=8, column=c + 3, value=income)
    if tax is not None:
        ws.cell(row=9, column=c, value="법인세 상당액")
        ws.cell(row=9, column=c + 3, value=tax)

    ws.cell(row=11, column=c, value="매출거래처")
    ws.cell(row=11, column=c + 1, value="매출액")
    ws.cell(row=11, column=c + 2, value="매출액 중 해외매출\n(L/C, 내국신용장 등)")
    ws.cell(row=11, column=c + 3, value="매출 합계 (1년 환산)")
    # 원본에 있는 설명 행. 거래처로 읽히면 안 된다.
    ws.cell(row=13, column=c + 1, value="발생한 실제매출")
    ws.cell(row=14, column=c + 1, value="A")
    ws.cell(row=14, column=c + 2, value="B")
    ws.cell(row=14, column=c + 3, value="(A-B)")

    r = 15
    for counterparty, actual, foreign, factor in rows:
        ws.cell(row=r, column=c - 2, value="CODE")
        ws.cell(row=r, column=c - 1, value="129-81-00178")
        ws.cell(row=r, column=c, value=counterparty)
        ws.cell(row=r, column=c + 1, value=actual)
        if foreign:
            ws.cell(row=r, column=c + 2, value=foreign)
        if isinstance(actual, (int, float)):
            ws.cell(row=r, column=c + 3, value=(actual - (foreign or 0)) * factor)
        r += 1
    # 표 아래 집계 행. 값이 비율(0.xx)이라 거래처로 읽으면 1원짜리 거래처가 생긴다.
    ws.cell(row=r + 1, column=c, value="특관매출 비율")
    ws.cell(row=r + 1, column=c + 1, value=0.9987)


def _workbook(build):
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    build(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse(content, fixture_data):
    calc = fixture_data
    return bulk_import.parse_workbook(content, "통합본.xlsx",
                                      calc.company_list(), calc.dataset().sizes)


def _by_sheet(parsed, title):
    return next(s for s in parsed["sheets"] if s["sheet"] == title)


# --- 파싱 ---------------------------------------------------------------------

def test_reads_sheets_with_different_column_offsets(fixture_data):
    """시트마다 표 시작 열이 달라도 같은 값을 읽어야 한다."""
    def build(wb):
        _sheet(wb.create_sheet("왼쪽"), name=SUBJECT, size="일반",
               total=10_000_000_000, income=1_000_000_000,
               rows=[(COUNTERPARTY, 500_000_000, 0, 1)], start_col=3)
        _sheet(wb.create_sheet("오른쪽"), name="다라화학", size="중견",
               total=10_000_000_000, income=1_000_000_000,
               rows=[(COUNTERPARTY, 500_000_000, 0, 1)], start_col=6)

    parsed = _parse(_workbook(build), fixture_data)
    left, right = _by_sheet(parsed, "왼쪽"), _by_sheet(parsed, "오른쪽")
    for s in (left, right):
        assert s["total_sales"] == 10_000_000_000
        assert s["operating_income"] == 1_000_000_000
        assert s["related_sales"] == {COUNTERPARTY: 500_000_000}
    assert left["company"] == SUBJECT
    assert right["company"] == "다라화학"


def test_uses_annualized_column_for_detail_rows(fixture_data):
    """거래처 매출은 파일의 연환산 열((A−B)×계수)을 그대로 쓴다.

    그 열이 이 파일의 정본이다 — 해외매출이 이미 빠져 있고 연환산도 끝나 있어,
    우리가 계수를 역산해 실매출에 곱하는 것보다 정확하다.
    총매출·영업이익은 파일 안에서 이미 환산돼 있으므로 손대지 않는다.
    """
    def build(wb):
        _sheet(wb.create_sheet("반기"), name=SUBJECT, size="일반",
               total=10_000_000_000, income=1_000_000_000,
               rows=[(COUNTERPARTY, 1_000_000_000, 0, 2),
                     (OTHER, 500_000_000, 100_000_000, 2)])

    s = _by_sheet(_parse(_workbook(build), fixture_data), "반기")
    assert s["related_sales"][COUNTERPARTY] == 2_000_000_000      # (10억 − 0) × 2
    assert s["related_sales"][OTHER] == 800_000_000               # (5억 − 1억) × 2
    # 해외매출은 연환산 열에서 이미 빠졌으므로 ⑩ 으로 다시 넘기지 않는다.
    # 두 번 빼면 판정비율이 실제보다 낮아진다.
    assert s["article10_exclusions"] == {}
    # 다만 얼마가 빠진 값인지는 남겨서 화면에 보여준다.
    assert s["foreign_sales_total"] == 200_000_000
    assert any("해외매출" in w for w in s["warnings"])
    assert s["total_sales"] == 10_000_000_000


def test_reads_value_past_placeholder_and_half_year_columns(fixture_data):
    """라벨과 값 사이에 안내 문구·반기 금액이 끼어 있어도 연환산 값을 읽는다.

    실제 파일이 이렇다:
        C6=총매출액 | D6='반기매출입력' | E6=454,217,451(반기) | F6=908,434,902(연환산)
    라벨 오른쪽 '첫' 칸을 값으로 삼으면 안내 문구를 읽고 입력대기로 빼버린다 —
    데이터가 있는 법인이 통째로 판정에서 빠지는 사고다(아이엔·대웅테라가 그랬다).
    """
    def build(wb):
        ws = wb.create_sheet("문구끼임")
        _sheet(ws, name=SUBJECT, size="일반", total=10_000_000_000, income=1_000_000_000,
               rows=[(COUNTERPARTY, 500_000_000, 0, 2)])
        # 라벨(D열)과 연환산 값(G열) 사이에 안내 문구와 반기 금액을 끼워 넣는다.
        ws.cell(row=6, column=5, value="반기매출입력")
        ws.cell(row=6, column=6, value=5_000_000_000)
        ws.cell(row=8, column=5, value="반기영업이익 입력")
        ws.cell(row=8, column=6, value=500_000_000)

    s = _by_sheet(_parse(_workbook(build), fixture_data), "문구끼임")
    assert s["status"] == "ok"
    assert s["total_sales"] == 10_000_000_000       # 반기 50억이 아니라 연환산 100억
    assert s["operating_income"] == 1_000_000_000


def test_placeholder_text_is_pending_not_zero(fixture_data):
    """'반기매출입력' 같은 안내 문구를 0 으로 읽어 '해당없음'을 내면 안 된다."""
    def build(wb):
        _sheet(wb.create_sheet("미입력"), name=SUBJECT, size="일반",
               total="반기매출입력", income="반기영업이익 입력",
               rows=[(COUNTERPARTY, 500_000_000, 0, 2)])

    s = _by_sheet(_parse(_workbook(build), fixture_data), "미입력")
    assert s["status"] == "입력대기"
    assert s["total_sales"] is None and s["operating_income"] is None
    assert any("반기매출입력" in w for w in s["warnings"])


def test_related_sales_over_total_sales_flagged(fixture_data):
    """총매출이 아직 임시값이면 비율이 터무니없어진다. 조용히 계산하지 않는다."""
    def build(wb):
        _sheet(wb.create_sheet("임시값"), name=SUBJECT, size="일반",
               total=2, income=2,
               rows=[(COUNTERPARTY, 500_000_000, 0, 2)])

    s = _by_sheet(_parse(_workbook(build), fixture_data), "임시값")
    assert s["status"] == "확인필요"
    assert any("총매출" in w for w in s["warnings"])


def test_placeholder_beats_zero_when_deciding_pending(fixture_data):
    """총매출이 '숫자 0' 인 것과 '안 적힌 것' 은 다르게 다뤄야 한다.

    0 은 매출이 없다는 사실이지만, 안내 문구는 아직 안 채운 것이다.
    전자는 판정 대상이 아니고(매출없음), 후자는 사람이 채워야 한다(입력대기).
    """
    def build(wb):
        _sheet(wb.create_sheet("영"), name=SUBJECT, size="일반", total=0, income=0, rows=[])
        _sheet(wb.create_sheet("미입력"), name="다라화학", size="중견",
               total="반기매출입력", income="반기영업이익 입력", rows=[])

    parsed = _parse(_workbook(build), fixture_data)
    assert _by_sheet(parsed, "영")["status"] == "매출없음"
    assert _by_sheet(parsed, "미입력")["status"] == "입력대기"


def test_size_mismatch_warns_and_keeps_server_value(fixture_data):
    """규모가 바뀌면 정상거래비율이 통째로 달라진다. 엑셀 표기를 따라가면 안 된다."""
    def build(wb):
        _sheet(wb.create_sheet("규모"), name=SUBJECT, size="대기업",
               total=10_000_000_000, income=1_000_000_000,
               rows=[(COUNTERPARTY, 500_000_000, 0, 1)])

    s = _by_sheet(_parse(_workbook(build), fixture_data), "규모")
    assert s["size_mismatch"] is True
    assert s["size_excel"] == "대기업"
    assert s["size_app"] == "일반"          # 서버 값이 정본
    assert any("기업구분" in w for w in s["warnings"])


def test_size_suffix_variants_are_not_mismatch(fixture_data):
    """'일반' 과 '일반기업' 은 같은 뜻이다. 매번 경고를 띄우면 진짜 불일치가 묻힌다."""
    def build(wb):
        _sheet(wb.create_sheet("표기"), name=SUBJECT, size="일반기업",
               total=10_000_000_000, income=1_000_000_000, rows=[])

    assert _by_sheet(_parse(_workbook(build), fixture_data), "표기")["size_mismatch"] is False


def test_sheet_name_abbreviation_resolved_by_company_cell(fixture_data):
    """시트명은 '생명과학' 처럼 줄임말이 흔하다. 법인명 셀이 정본이다."""
    def build(wb):
        _sheet(wb.create_sheet("가나"), name=SUBJECT, size="일반",
               total=10_000_000_000, income=1_000_000_000, rows=[])

    s = _by_sheet(_parse(_workbook(build), fixture_data), "가나")
    assert s["company"] == SUBJECT
    assert s["status"] == "ok"


def test_summary_row_is_not_read_as_counterparty(fixture_data):
    """표 아래 '특관매출 비율' 행의 0.9987 이 1원짜리 거래처가 되면 안 된다."""
    def build(wb):
        _sheet(wb.create_sheet("집계"), name=SUBJECT, size="일반",
               total=10_000_000_000, income=1_000_000_000,
               rows=[(COUNTERPARTY, 500_000_000, 0, 1)])

    s = _by_sheet(_parse(_workbook(build), fixture_data), "집계")
    assert list(s["related_sales"]) == [COUNTERPARTY]
    assert s["unmatched"] == []


def test_unknown_counterparty_is_reported_not_dropped(fixture_data):
    def build(wb):
        _sheet(wb.create_sheet("낯선"), name=SUBJECT, size="일반",
               total=10_000_000_000, income=1_000_000_000,
               rows=[("없는거래처", 300_000_000, 0, 1)])

    s = _by_sheet(_parse(_workbook(build), fixture_data), "낯선")
    assert [u["name"] for u in s["unmatched"]] == ["없는거래처"]
    assert s["related_sales"] == {}


def test_corporate_tax_read_when_present_else_warned(fixture_data):
    """지금 통합본에는 법인세가 없다. 나중에 양식에 생기면 자동으로 읽혀야 한다."""
    def build(wb):
        _sheet(wb.create_sheet("있음"), name=SUBJECT, size="일반",
               total=10_000_000_000, income=1_000_000_000, rows=[], tax=200_000_000)
        _sheet(wb.create_sheet("없음"), name="다라화학", size="중견",
               total=10_000_000_000, income=1_000_000_000, rows=[])

    parsed = _parse(_workbook(build), fixture_data)
    assert _by_sheet(parsed, "있음")["corporate_tax"] == 200_000_000
    absent = _by_sheet(parsed, "없음")
    assert absent["corporate_tax"] is None
    assert any("법인세" in w for w in absent["warnings"])


def test_missing_companies_listed(fixture_data):
    """시트가 아예 없는 법인은 '미제출'로 드러나야 한다. 빠진 줄 모르는 것이 문제다."""
    def build(wb):
        _sheet(wb.create_sheet("하나"), name=SUBJECT, size="일반",
               total=10_000_000_000, income=1_000_000_000, rows=[])

    parsed = _parse(_workbook(build), fixture_data)
    assert SUBJECT not in parsed["missing_companies"]
    assert "다라화학" in parsed["missing_companies"]
    assert parsed["stats"]["missing"] == len(parsed["missing_companies"])
    # 기타법인은 거래처일 뿐 판정 대상이 아니다.
    assert fixture_data.OTHER_COMPANY not in parsed["missing_companies"]


def test_rejects_non_excel(fixture_data):
    with pytest.raises(ValueError):
        _parse(b"not an excel file", fixture_data)


# --- 엔드포인트 ---------------------------------------------------------------

def _admin():
    r = client.post("/api/auth/login",
                    json={"username": ADMIN_USERNAME, "password": ADMIN_PASSWORD})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def test_bulk_parse_requires_admin(fixture_data):
    files = {"file": ("x.xlsx", b"PK", "application/vnd.ms-excel")}
    assert client.post("/api/admin/bulk/parse", files=files).status_code == 401


def test_bulk_evaluate_requires_admin(fixture_data):
    body = {"companies": [{"company": SUBJECT, "operating_income": 0, "total_sales": 0}]}
    assert client.post("/api/admin/bulk/evaluate", json=body).status_code == 401


def test_bulk_evaluate_totals_match_per_company(fixture_data):
    body = {"companies": [
        {"company": SUBJECT, "operating_income": 10_000_000_000, "corporate_tax": 0,
         "total_sales": 10_000_000_000, "related_sales": {COUNTERPARTY: 5_000_000_000}},
        {"company": "다라화학", "operating_income": 1_000_000_000, "corporate_tax": 0,
         "total_sales": 10_000_000_000, "related_sales": {COUNTERPARTY: 100_000_000}},
    ]}
    r = client.post("/api/admin/bulk/evaluate", json=body, headers=_admin())
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["failed"] == []
    assert data["totals"]["evaluated"] == 2
    assert data["totals"]["gift_tax_total"] == sum(x["gift_tax_total"] for x in data["results"])
    assert data["totals"]["taxable_count"] == sum(1 for x in data["results"] if x["taxable"])


def test_bulk_evaluate_reports_bad_company_without_failing_the_rest(fixture_data):
    """한 법인이 틀렸다고 나머지 판정까지 버리면 15개짜리 통합본을 못 쓴다."""
    body = {"companies": [
        {"company": "없는법인", "operating_income": 0, "total_sales": 1},
        {"company": SUBJECT, "operating_income": 10_000_000_000, "corporate_tax": 0,
         "total_sales": 10_000_000_000, "related_sales": {COUNTERPARTY: 5_000_000_000}},
    ]}
    data = client.post("/api/admin/bulk/evaluate", json=body, headers=_admin()).json()
    assert [f["company"] for f in data["failed"]] == ["없는법인"]
    assert [x["company"] for x in data["results"]] == [SUBJECT]


def test_bulk_evaluate_rejects_unknown_year(fixture_data):
    body = {"year": "1999",
            "companies": [{"company": SUBJECT, "operating_income": 0, "total_sales": 0}]}
    assert client.post("/api/admin/bulk/evaluate", json=body,
                       headers=_admin()).status_code == 400


def test_non_company_sheets_are_skipped_not_flagged(fixture_data):
    """요약·분류 시트는 '미매칭'이 아니라 '건너뜀'이다.

    실제 통합본에는 '특관매출 요약', '기업분류', '0. 보고' 같은 시트가 섞여 있다.
    이걸 미매칭으로 세면 진짜 확인이 필요한 법인이 그 안에 묻힌다.
    """
    def build(wb):
        _sheet(wb.create_sheet("가나"), name=SUBJECT, size="일반",
               total=10_000_000_000, income=1_000_000_000, rows=[])
        summary = wb.create_sheet("특관매출 요약")
        summary["B2"] = "특관매출 요약"
        summary["B4"] = "법인별 집계"
        summary["C4"] = 12345

    parsed = _parse(_workbook(build), fixture_data)
    skipped = _by_sheet(parsed, "특관매출 요약")
    assert skipped["status"] == "건너뜀"
    assert parsed["stats"]["skipped"] == 1
    # 보류로 세면 안 된다 — 사람이 손볼 것이 없는 시트다.
    assert parsed["stats"]["pending"] == 0
    assert parsed["stats"]["ready"] == 1


def test_blank_size_cell_is_not_a_mismatch(fixture_data):
    """기업구분 칸이 0/빈칸이면 '다르다'가 아니라 '안 적혔다'다.

    매번 경고를 띄우면 진짜 불일치(엑셀 '대기업' vs 서버 '일반')가 묻힌다.
    """
    def build(wb):
        _sheet(wb.create_sheet("빈칸"), name=SUBJECT, size=0,
               total=10_000_000_000, income=1_000_000_000, rows=[])

    s = _by_sheet(_parse(_workbook(build), fixture_data), "빈칸")
    assert s["size_mismatch"] is False
    assert not any("기업구분" in w for w in s["warnings"])
    assert s["size_app"] == "일반"      # 계산은 서버 값으로 한다


def test_zero_everything_is_no_sales_not_pending(fixture_data):
    """총매출도 특관매출도 0 이면 판정 대상이 아니다.

    실제로 매출이 없는 법인이 있다(대웅낙원·블루넷 등). 이런 곳을 '보류'로 세면
    사람이 손봐야 할 법인이 그 안에 묻힌다.
    """
    def build(wb):
        _sheet(wb.create_sheet("무매출"), name=SUBJECT, size="일반", total=0, income=-1_000_000,
               rows=[])
        _sheet(wb.create_sheet("정상"), name="다라화학", size="중견",
               total=10_000_000_000, income=1_000_000_000,
               rows=[(COUNTERPARTY, 500_000_000, 0, 1)])

    parsed = _parse(_workbook(build), fixture_data)
    s = _by_sheet(parsed, "무매출")
    assert s["status"] == "매출없음"
    assert parsed["stats"]["no_sales"] == 1
    assert parsed["stats"]["pending"] == 0      # 보류로 세지 않는다
    assert parsed["stats"]["ready"] == 1


def test_zero_total_with_related_sales_is_flagged(fixture_data):
    """총매출만 0 이고 특관매출이 있으면 그냥 넘기면 안 된다.

    '매출이 없는 법인'이 아니라 데이터가 어긋난 것이다. 비율이 성립하지 않으므로
    조용히 제외하면 과세 대상이 통째로 빠질 수 있다.
    """
    def build(wb):
        _sheet(wb.create_sheet("모순"), name=SUBJECT, size="일반", total=0, income=1_000_000_000,
               rows=[(COUNTERPARTY, 500_000_000, 0, 1)])

    parsed = _parse(_workbook(build), fixture_data)
    s = _by_sheet(parsed, "모순")
    assert s["status"] == "확인필요"
    assert parsed["stats"]["no_sales"] == 0
    assert parsed["stats"]["pending"] == 1
    assert any("총매출이 0" in w for w in s["warnings"])


def test_long_instruction_text_is_not_read_as_a_label(fixture_data):
    """상세표 옆 안내 문단을 라벨로 잡으면 안 된다.

    실제 통합본에는 'List에 없는 특수관계자 매출 내역은 ... (법인세 서식 52호 기준)'
    이라는 문단이 있다. '법인세' 가 그 안에 들어 있어 부분일치로 법인세 라벨이 되고,
    그 오른쪽 아무 숫자나 법인세 상당액으로 읽힐 수 있다.
    """
    def build(wb):
        ws = wb.create_sheet("안내문")
        _sheet(ws, name=SUBJECT, size="일반", total=10_000_000_000, income=1_000_000_000,
               rows=[(COUNTERPARTY, 500_000_000, 0, 1)])
        ws.cell(row=11, column=9,
                value="List에 없는 특수관계자 매출 내역은 특수관계자 추가해서 "
                      "작성 부탁드립니다. (법인세 서식 52호 기준)")
        ws.cell(row=11, column=10, value=999_999_999)

    s = _by_sheet(_parse(_workbook(build), fixture_data), "안내문")
    assert s["corporate_tax"] is None, "안내 문단 옆 숫자를 법인세로 읽었다"
    assert any("법인세" in w for w in s["warnings"])


def test_tax_adjustment_read_when_present(fixture_data):
    """세무조정은 지금 양식에 없다. 나중에 생기면 코드 수정 없이 읽혀야 한다."""
    def build(wb):
        ws = wb.create_sheet("조정")
        _sheet(ws, name=SUBJECT, size="일반", total=10_000_000_000, income=1_000_000_000,
               rows=[], tax=200_000_000)
        # 차감 조정은 음수로 적힌다.
        ws.cell(row=7, column=4, value="세무조정 합계")
        ws.cell(row=7, column=7, value=-3_500_000)

    parsed = _parse(_workbook(build), fixture_data)
    assert _by_sheet(parsed, "조정")["tax_adjustment"] == -3_500_000
    # 없는 시트는 None 이고, 화면에서 입력받는다.
    def build2(wb):
        _sheet(wb.create_sheet("없음"), name=SUBJECT, size="일반",
               total=10_000_000_000, income=1_000_000_000, rows=[])
    assert _by_sheet(_parse(_workbook(build2), fixture_data), "없음")["tax_adjustment"] is None


# --- 빈 시트 양식 ---------------------------------------------------------------

def test_blank_sheets_round_trip(fixture_data):
    """만들어 준 빈 시트를 채워서 다시 올리면 그대로 읽혀야 한다.

    양식과 파서가 어긋나면 사용자는 채워 넣고도 '입력대기'만 보게 된다.
    """
    import openpyxl

    calc = fixture_data
    companies = [SUBJECT, "다라화학"]
    data = bulk_import.build_blank_sheets(companies, calc.company_list(),
                                          calc.dataset().sizes, "테스트 기준")

    # 안 채운 채로 올려도 조용히 0 으로 계산되면 안 된다.
    blank = bulk_import.parse_workbook(data, "blank.xlsx", calc.company_list(),
                                       calc.dataset().sizes)
    assert [s["sheet"] for s in blank["sheets"]] == companies
    assert {s["status"] for s in blank["sheets"]} == {"입력대기"}

    # 엑셀에서 채워 저장한 상태를 흉내 낸다(엑셀은 수식 결과를 캐시해 둔다).
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb[SUBJECT]
    ws["E6"], ws["F6"] = 5_000_000_000, 10_000_000_000     # 반기 → 연환산
    ws["E8"], ws["F8"] = 400_000_000, 800_000_000
    for row in ws.iter_rows(min_row=bulk_import.BLANK_FIRST_ROW, max_col=6):
        if row[2].value == COUNTERPARTY:
            row[3].value, row[4].value, row[5].value = 1_000_000_000, 0, 2_000_000_000
        elif isinstance(row[5].value, str) and row[5].value.startswith("="):
            row[5].value = 0
    buf = io.BytesIO()
    wb.save(buf)

    filled = bulk_import.parse_workbook(buf.getvalue(), "filled.xlsx",
                                        calc.company_list(), calc.dataset().sizes)
    s = next(x for x in filled["sheets"] if x["sheet"] == SUBJECT)
    assert s["status"] == "ok"
    assert s["total_sales"] == 10_000_000_000       # 반기 50억이 아니라 연환산 100억
    assert s["operating_income"] == 800_000_000
    assert s["related_sales"] == {COUNTERPARTY: 2_000_000_000}


def test_blank_sheets_use_server_size_and_company_list(fixture_data):
    """기업구분은 서버 값을 박아 주고, 거래처 목록도 서버 목록을 그대로 쓴다."""
    import openpyxl

    calc = fixture_data
    data = bulk_import.build_blank_sheets([SUBJECT], calc.company_list(),
                                          calc.dataset().sizes)
    ws = openpyxl.load_workbook(io.BytesIO(data))[SUBJECT]
    assert ws["F2"].value == SUBJECT
    assert ws["F4"].value == calc.dataset().sizes[SUBJECT]
    names = [ws.cell(row=r, column=3).value
             for r in range(bulk_import.BLANK_FIRST_ROW,
                            bulk_import.BLANK_FIRST_ROW + len(calc.company_list()))]
    assert names == calc.company_list()


def test_blank_sheets_endpoint_requires_admin_and_known_company(fixture_data):
    body = {"companies": [SUBJECT]}
    assert client.post("/api/admin/bulk/blank-sheets", json=body).status_code == 401

    r = client.post("/api/admin/bulk/blank-sheets", json=body, headers=_admin())
    assert r.status_code == 200, r.text
    assert r.content[:2] == b"PK"          # xlsx 는 zip 이다

    bad = client.post("/api/admin/bulk/blank-sheets",
                      json={"companies": ["없는법인"]}, headers=_admin())
    assert bad.status_code == 400
