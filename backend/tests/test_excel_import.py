"""특수관계자 매출 엑셀 업로드 파싱 테스트.

핵심 불변조건은 두 가지다.
  - 우리가 내려준 양식은 왕복(다운로드 → 금액 채움 → 업로드)이 정확히 맞아떨어져야 한다.
  - 확신이 없으면 매칭하지 않는다. 금액이 엉뚱한 법인에 붙는 것이 최악이다.
"""

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from backend import excel_import
from backend.main import app

COMPANIES = ["대웅", "대웅제약", "대웅바이오", "이지메디컴", "기타법인"]


def _xlsx_bytes(rows):
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- 이름 정규화 -------------------------------------------------------------

@pytest.mark.parametrize("raw", [
    "대웅제약", "(주)대웅제약", "㈜대웅제약", "대웅제약(주)", "주식회사 대웅제약",
    " 대웅 제약 ", "대웅제약 (서울)", "대웅제약(주)  ",
])
def test_normalize_name_collapses_legal_forms(raw):
    assert excel_import.normalize_name(raw) == "대웅제약"


def test_normalize_name_keeps_distinct_companies_distinct():
    # 법인격만 걷어내야지, 서로 다른 회사를 같게 만들면 안 된다.
    assert excel_import.normalize_name("대웅제약") != excel_import.normalize_name("대웅바이오")


def test_normalize_name_handles_empty():
    assert excel_import.normalize_name(None) == ""
    assert excel_import.normalize_name("   ") == ""


# --- 금액 파싱 ---------------------------------------------------------------

@pytest.mark.parametrize("raw,expected", [
    (1200000, 1200000),
    (1200000.0, 1200000),
    ("1200000", 1200000),
    ("1,200,000", 1200000),
    ("1,200,000원", 1200000),
    ("₩1,200,000", 1200000),
    ("(500,000)", -500000),        # 회계식 음수
    ("-500,000", -500000),
    ("0", 0),
])
def test_parse_amount(raw, expected):
    assert excel_import.parse_amount(raw) == expected


@pytest.mark.parametrize("raw", [None, "", "   ", "미정", "N/A", "-", True, False])
def test_parse_amount_rejects_non_numbers(raw):
    assert excel_import.parse_amount(raw) is None


# --- 양식 왕복 ---------------------------------------------------------------

def test_template_roundtrip_matches_every_company():
    """다운로드한 양식에 금액만 채워 올리면 전부 매칭돼야 한다. 미매칭이 하나라도 나오면 안 된다."""
    template = excel_import.build_template(COMPANIES)
    wb = load_workbook(io.BytesIO(template))
    ws = wb.active

    assert [c.value for c in ws[1]] == ["거래처명", "매출액(원)"]
    names = [ws.cell(row=i, column=1).value for i in range(2, 2 + len(COMPANIES))]
    assert names == COMPANIES

    for offset in range(len(COMPANIES)):
        ws.cell(row=2 + offset, column=2, value=(offset + 1) * 100000)
    buf = io.BytesIO()
    wb.save(buf)

    result = excel_import.import_related_sales(buf.getvalue(), "filled.xlsx", COMPANIES)
    assert result["unmatched"] == []
    assert result["stats"]["matched_count"] == len(COMPANIES)
    amounts = {m["company"]: m["amount"] for m in result["matched"]}
    assert amounts == {c: (i + 1) * 100000 for i, c in enumerate(COMPANIES)}


def test_template_note_row_is_not_read_as_a_counterparty():
    """양식 하단 안내문이 거래처로 잡히면 안 된다(금액이 없으므로 걸러져야 한다)."""
    template = excel_import.build_template(COMPANIES)
    wb = load_workbook(io.BytesIO(template))
    ws = wb.active
    ws.cell(row=2, column=2, value=100)
    buf = io.BytesIO()
    wb.save(buf)
    result = excel_import.import_related_sales(buf.getvalue(), "t.xlsx", COMPANIES)
    assert result["unmatched"] == []


# --- ERP 추출본 --------------------------------------------------------------

ERP_ROWS = [
    ["[매출처별 집계표]", None, None, None],
    ["기간: 2025.01 ~ 2025.12", None, None, None],
    [None, None, None, None],
    ["거래처코드", "거래처명", "매출액", "비고"],
    ["1001", "(주)대웅제약", "1,200,000", ""],
    ["1002", "대웅바이오(주)", "850,000", ""],
    ["1003", "한올바이오파마", "300,000", "특수관계"],
    ["", "합계", "2,350,000", ""],
]


def test_erp_export_finds_header_below_title_rows():
    result = excel_import.import_related_sales(_xlsx_bytes(ERP_ROWS), "erp.xlsx", COMPANIES)
    amounts = {m["company"]: m["amount"] for m in result["matched"]}
    assert amounts == {"대웅제약": 1200000, "대웅바이오": 850000}
    # 목록에 없는 거래처는 조용히 버리지 않고 그대로 돌려준다.
    assert result["unmatched"] == [{"name": "한올바이오파마", "amount": 300000}]


def test_erp_export_skips_total_row():
    """합계 행을 거래처로 잡으면 매출이 두 배가 된다."""
    result = excel_import.import_related_sales(_xlsx_bytes(ERP_ROWS), "erp.xlsx", COMPANIES)
    total = result["stats"]["matched_total"] + result["stats"]["unmatched_total"]
    assert total == 2350000
    assert all("합계" not in m["company"] for m in result["matched"])
    assert all("합계" not in u["name"] for u in result["unmatched"])


def test_code_column_is_not_used_as_name():
    """'거래처코드'가 '거래처'에 걸려 이름 열로 뽑히면 전부 미매칭이 된다."""
    result = excel_import.import_related_sales(_xlsx_bytes(ERP_ROWS), "erp.xlsx", COMPANIES)
    assert result["stats"]["matched_count"] == 2


def test_duplicate_rows_are_summed():
    """지점·월별로 쪼개진 파일은 같은 법인이 여러 줄에 나온다."""
    rows = [
        ["거래처명", "매출액"],
        ["(주)대웅제약", 100],
        ["대웅제약", 200],
        ["㈜대웅제약", 300],
    ]
    result = excel_import.import_related_sales(_xlsx_bytes(rows), "dup.xlsx", COMPANIES)
    assert len(result["matched"]) == 1
    assert result["matched"][0]["company"] == "대웅제약"
    assert result["matched"][0]["amount"] == 600
    assert len(result["matched"][0]["sources"]) == 3


def test_unmatched_rows_are_summed_too():
    rows = [
        ["거래처명", "매출액"],
        ["한올바이오파마", 100],
        ["(주)한올바이오파마", 200],
    ]
    result = excel_import.import_related_sales(_xlsx_bytes(rows), "u.xlsx", COMPANIES)
    assert result["unmatched"] == [{"name": "한올바이오파마", "amount": 300}]


# --- 이름 별칭 ---------------------------------------------------------------

def test_other_company_constant_matches_calc():
    """excel_import 가 상수를 따로 들고 있으므로 calc 와 어긋나면 여기서 잡는다."""
    from backend import calc
    assert excel_import.OTHER_COMPANY == calc.OTHER_COMPANY


@pytest.mark.parametrize("alias", ["기타", "기타 ", "기타거래처", "기타매출처", "그 외", "기타업체"])
def test_catch_all_aliases_map_to_other_company(alias):
    """파일에는 '기타'로, 서버 목록에는 '기타법인'으로 적혀 매칭이 안 되던 건."""
    rows = [["거래처명", "매출액"], [alias, 500000]]
    result = excel_import.import_related_sales(_xlsx_bytes(rows), "a.xlsx", COMPANIES)
    assert result["unmatched"] == []
    assert result["matched"] == [
        {"company": "기타법인", "amount": 500000, "sources": [alias.strip()]}
    ]


def test_alias_and_canonical_rows_are_summed():
    """'기타'와 '기타법인'이 한 파일에 같이 있으면 합쳐야 한다."""
    rows = [["거래처명", "매출액"], ["기타", 100], ["기타법인", 200], ["기타거래처", 300]]
    result = excel_import.import_related_sales(_xlsx_bytes(rows), "a.xlsx", COMPANIES)
    assert result["matched"] == [
        {"company": "기타법인", "amount": 600, "sources": ["기타", "기타법인", "기타거래처"]}
    ]


def test_alias_never_overrides_a_real_company_name():
    """별칭이 실제 법인명을 가로채면 안 된다."""
    companies = ["기타", "기타법인"]     # '기타'라는 법인이 실제로 있는 극단적 경우
    rows = [["거래처명", "매출액"], ["기타", 100]]
    result = excel_import.import_related_sales(_xlsx_bytes(rows), "a.xlsx", companies)
    assert result["matched"][0]["company"] == "기타"


def test_alias_is_ignored_when_target_absent_from_list():
    """가리키는 법인이 목록에 없으면 별칭은 없는 셈 쳐야 한다(없는 법인을 만들지 않는다)."""
    companies = ["대웅제약"]            # 기타법인 없음
    rows = [["거래처명", "매출액"], ["기타", 100]]
    result = excel_import.import_related_sales(_xlsx_bytes(rows), "a.xlsx", companies)
    assert result["matched"] == []
    assert result["unmatched"] == [{"name": "기타", "amount": 100}]


def test_rows_without_amount_are_skipped_with_warning():
    rows = [
        ["거래처명", "매출액"],
        ["대웅제약", 100],
        ["대웅바이오", "미정"],
    ]
    result = excel_import.import_related_sales(_xlsx_bytes(rows), "s.xlsx", COMPANIES)
    assert result["stats"]["matched_count"] == 1
    assert any("건너뛰" in w for w in result["warnings"])


def test_negative_amount_is_flagged():
    rows = [["거래처명", "매출액"], ["대웅제약", "(50,000)"]]
    result = excel_import.import_related_sales(_xlsx_bytes(rows), "n.xlsx", COMPANIES)
    assert result["matched"][0]["amount"] == -50000
    assert any("음수" in w for w in result["warnings"])


def test_headerless_file_falls_back_to_sniffing_and_warns():
    rows = [["대웅제약", 100], ["대웅바이오", 200]]
    result = excel_import.import_related_sales(_xlsx_bytes(rows), "h.xlsx", COMPANIES)
    assert result["stats"]["matched_count"] == 2
    assert any("머리글" in w for w in result["warnings"])


# --- CSV ---------------------------------------------------------------------

def test_csv_utf8():
    content = "거래처명,매출액\n대웅제약,1200000\n".encode("utf-8")
    result = excel_import.import_related_sales(content, "a.csv", COMPANIES)
    assert result["matched"][0]["amount"] == 1200000


def test_csv_cp949():
    """국내 회계 프로그램은 CSV 를 cp949 로 뱉는 일이 흔하다."""
    content = "거래처명,매출액\n대웅제약,1200000\n".encode("cp949")
    result = excel_import.import_related_sales(content, "a.csv", COMPANIES)
    assert result["matched"][0]["amount"] == 1200000


def test_tab_separated():
    content = "거래처명\t매출액\n대웅제약\t1200000\n".encode("utf-8")
    result = excel_import.import_related_sales(content, "a.csv", COMPANIES)
    assert result["matched"][0]["amount"] == 1200000


# --- 거부해야 하는 입력 ------------------------------------------------------

def test_legacy_xls_gives_actionable_message():
    with pytest.raises(ValueError, match="xlsx"):
        excel_import.import_related_sales(b"\xd0\xcf\x11\xe0", "old.xls", COMPANIES)


def test_empty_file_rejected():
    with pytest.raises(ValueError, match="빈 파일"):
        excel_import.import_related_sales(b"", "a.xlsx", COMPANIES)


def test_oversized_file_rejected():
    big = b"x" * (excel_import.MAX_UPLOAD_BYTES + 1)
    with pytest.raises(ValueError, match="너무 큽니다"):
        excel_import.import_related_sales(big, "a.csv", COMPANIES)


def test_file_without_usable_rows_rejected():
    rows = [["안내문", "내용"], ["이 파일은", "매출 자료가 아닙니다"]]
    with pytest.raises(ValueError):
        excel_import.import_related_sales(_xlsx_bytes(rows), "x.xlsx", COMPANIES)


def test_row_cap_is_reported_not_silent():
    """잘라 읽었으면 반드시 알려야 한다. 조용히 자르면 누락을 눈치챌 수 없다."""
    rows = [["거래처명", "매출액"]]
    rows += [[f"회사{i}", 1] for i in range(excel_import.MAX_DATA_ROWS + 10)]
    result = excel_import.import_related_sales(_xlsx_bytes(rows), "big.xlsx", COMPANIES)
    assert any("행이 너무 많아" in w for w in result["warnings"])


# --- API ---------------------------------------------------------------------

client = TestClient(app)


def test_template_endpoint_requires_auth():
    assert client.get("/api/related-sales/template").status_code == 401


def test_import_endpoint_requires_auth():
    r = client.post("/api/related-sales/import", files={"file": ("a.csv", b"x", "text/csv")})
    assert r.status_code == 401
