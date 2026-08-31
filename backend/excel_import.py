"""특수관계자 세부매출 엑셀 업로드 파싱.

거래처별 매출을 화면에서 한 줄씩 입력하는 대신 파일로 채우기 위한 모듈이다.
두 가지 입력을 모두 받아야 한다.

1. `/api/related-sales/template` 로 내려준 양식 — 헤더가 1행, A열 거래처명, B열 매출액.
2. 회계/ERP 에서 뽑은 기존 파일 — 제목·기간 같은 머리글이 위에 붙고, 코드/비고 열이
   섞여 있으며, 거래처명 표기도 '(주)대웅제약' 처럼 서버 목록과 다르다.

그래서 헤더 행과 열 위치를 찾아내고, 이름을 정규화해 맞춰본다. **판단이 애매하면
맞췄다고 우기지 않고 미매칭으로 넘긴다** — 금액이 엉뚱한 법인에 붙는 것이 못 찾는 것보다
훨씬 나쁘다. 미매칭은 호출부가 사용자에게 보여주고 직접 연결하게 한다.

이 모듈은 순수 파싱만 한다. 인증·권한·법인 목록 확정은 main.py 가 맡는다.
"""

import csv
import io
import re
from typing import Dict, List, Optional, Tuple

# 업로드 상한. 실수로 수십 MB 짜리 원장을 통째로 올리는 것을 막는다.
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_DATA_ROWS = 5000
# 헤더는 파일 맨 위에 있다. 이보다 아래에서 찾으면 데이터 행을 헤더로 오인한다.
HEADER_SCAN_ROWS = 30
# 헤더를 못 찾았을 때 열을 추정하기 위해 훑는 행 수.
SNIFF_ROWS = 60

# 헤더 후보. 부분일치로 본다('거래처명', '매출처 코드' 등 변형이 많다).
NAME_HEADERS = ("거래처", "업체", "회사", "법인", "상호", "매출처", "고객", "거래상대")
AMOUNT_HEADERS = ("매출액", "매출", "금액", "공급가액", "합계", "총액", "amount", "sales")
# ⑩ 과세제외액 열(선택). 양식에는 늘 있지만 ERP 파일에는 대개 없다 — 없으면 0 으로 본다.
# `과세제외` 가 들어간 머리글은 금액 열 후보이기도 해서, 금액 열보다 **먼저** 집어낸다.
ARTICLE10_HEADERS = ("과세제외", "제10항", "10항", "⑩")
# 이름 열로 쓰면 안 되는 것들. '거래처코드'가 '거래처'에 걸리는 것을 막는다.
NAME_HEADER_BLOCKERS = ("코드", "번호", "code", "no.", "사업자")

# 합계 행을 거래처로 잡으면 매출이 두 배가 된다.
TOTAL_ROW_MARKERS = ("합계", "소계", "총계", "계", "total", "sum", "누계")

# 법인격 표기. 같은 회사가 파일마다 다르게 적히므로 비교 전에 걷어낸다.
_LEGAL_FORMS = (
    "주식회사", "유한회사", "유한책임회사", "합자회사", "합명회사",
    "재단법인", "사단법인", "의료법인", "학교법인",
    "co.,ltd", "co.ltd", "coltd", "ltd", "inc", "corp", "corporation", "company",
)
_LEGAL_MARKS = "㈜㈎㈏㈐㈑㈒㈓㈔"

# 서버 목록의 catch-all 거래처. calc.OTHER_COMPANY 와 같은 문자열이어야 하며,
# test_excel_import.py 가 어긋나면 실패시킨다(여기서 calc 를 import 하면 이 모듈이
# 지분 데이터 적재에 묶여버리므로 상수를 두고 테스트로 지킨다).
OTHER_COMPANY = "기타법인"

# 파일에는 있지만 서버 목록과 표기가 다른 이름을 이어준다.
# 키는 normalize_name() 을 거친 형태로 적고, 값은 서버 법인 목록에 있는 이름이어야 한다.
# 목록에 없으면 그 별칭은 그냥 무시된다 — 별칭 때문에 없는 법인이 생기면 안 된다.
#
# 법인격 표기 차이('(주)대웅제약')는 normalize_name 이 이미 처리하므로 여기 적지 않는다.
# 여기는 **철자가 아예 다른** 경우만 넣는다.
NAME_ALIASES = {
    "기타": OTHER_COMPANY,
    "기타거래처": OTHER_COMPANY,
    "기타매출처": OTHER_COMPANY,
    "기타업체": OTHER_COMPANY,
    "기타법인등": OTHER_COMPANY,
    "그외": OTHER_COMPANY,
    "그밖의법인": OTHER_COMPANY,
}


def normalize_name(value) -> str:
    """비교용 이름. 법인격·괄호·공백·기호를 없앤 뒤 소문자로 만든다.

    '(주) 대웅 제약', '주식회사대웅제약', '대웅제약(주)' 이 모두 '대웅제약'이 되게 하는 것이 목적이다.
    """
    if value is None:
        return ""
    s = str(value)
    for mark in _LEGAL_MARKS:
        s = s.replace(mark, "")
    # '(주)' 처럼 괄호에 싸인 법인격은 괄호째 지운다. 지점 표기 '(서울)' 도 같이 지워지는데
    # 그건 의도한 것이다 — 지점명이 붙어도 본사 이름으로 맞춰야 한다.
    s = re.sub(r"[\(\[\{][^\)\]\}]{0,12}[\)\]\}]", "", s)
    lowered = s.lower()
    for form in _LEGAL_FORMS:
        lowered = lowered.replace(form, "")
    # 남는 것은 한글/영문/숫자뿐이다.
    return re.sub(r"[^0-9a-z가-힣]", "", lowered)


def parse_amount(value) -> Optional[int]:
    """셀 값을 원 단위 정수로. 숫자로 볼 수 없으면 None.

    '1,200,000', '1200000원', '(500,000)'(회계식 음수), 1200000.0 을 모두 받는다.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(round(value))
    s = str(value).strip()
    if not s:
        return None
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1]
    s = s.replace(",", "").replace("원", "").replace("₩", "").strip()
    if s.startswith("-"):
        negative = True
        s = s[1:]
    if not re.fullmatch(r"\d+(\.\d+)?", s):
        return None
    amount = int(round(float(s)))
    return -amount if negative else amount


def _is_total_row(cells) -> bool:
    for cell in cells:
        if cell is None:
            continue
        text = re.sub(r"\s", "", str(cell))
        if text and text in TOTAL_ROW_MARKERS:
            return True
    return False


def _header_score(cells) -> Tuple[int, Optional[int], Optional[int]]:
    """행이 헤더처럼 보이는지 점수화하고 (이름열, 금액열) 을 돌려준다.

    제목 행 하나가 통째로 헤더로 오인되는 것을 막는 것이 이 함수의 핵심이다.
    '[매출처별 집계표]' 같은 제목은 '매출처'(이름)와 '매출'(금액)에 동시에 걸리는데,
    실제 헤더라면 이름 칸과 금액 칸이 **서로 다른 열**에 있다. 그래서 두 조건을 건다.
      - 셀이 2개 이상 채워져 있을 것 (제목 행은 보통 한 칸짜리다)
      - 이름 열과 금액 열이 다를 것
    """
    filled = [c for c in cells if c is not None and str(c).strip()]
    if len(filled) < 2:
        return 0, None, None

    name_cols: List[int] = []
    amount_cols: List[int] = []
    for idx, cell in enumerate(cells):
        if cell is None:
            continue
        text = str(cell).strip().lower()
        if not text or len(text) > 30:
            continue
        if any(h in text for h in NAME_HEADERS) and not any(b in text for b in NAME_HEADER_BLOCKERS):
            name_cols.append(idx)
        if any(h in text for h in AMOUNT_HEADERS):
            amount_cols.append(idx)

    name_col = name_cols[0] if name_cols else None
    # 금액 열은 이름 열과 겹칠 수 없다. 겹치는 후보는 건너뛰고 다음 후보를 본다.
    amount_col = next((i for i in amount_cols if i != name_col), None)
    if name_col is None or amount_col is None:
        return 0, None, None
    return 2, name_col, amount_col


def _sniff_columns(rows: List[list]) -> Tuple[Optional[int], Optional[int]]:
    """헤더가 없는 파일용. 글자가 가장 많은 열을 이름, 숫자가 가장 많은 열을 금액으로 본다."""
    text_hits: Dict[int, int] = {}
    num_hits: Dict[int, int] = {}
    for cells in rows[:SNIFF_ROWS]:
        for idx, cell in enumerate(cells):
            if cell is None or str(cell).strip() == "":
                continue
            if parse_amount(cell) is not None:
                num_hits[idx] = num_hits.get(idx, 0) + 1
            elif normalize_name(cell):
                text_hits[idx] = text_hits.get(idx, 0) + 1
    name_col = max(text_hits, key=lambda k: (text_hits[k], -k)) if text_hits else None
    # 금액 열이 이름 열과 겹치면 안 된다.
    candidates = {k: v for k, v in num_hits.items() if k != name_col}
    amount_col = max(candidates, key=lambda k: (candidates[k], -k)) if candidates else None
    return name_col, amount_col


def read_rows(content: bytes, filename: str) -> List[list]:
    """업로드 바이트를 행 리스트로. 지원하지 않는 형식이면 ValueError."""
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError(
            f"파일이 너무 큽니다. {MAX_UPLOAD_BYTES // (1024 * 1024)}MB 이하로 올려주세요."
        )
    if not content:
        raise ValueError("빈 파일입니다.")

    lower = (filename or "").lower()
    if lower.endswith(".xls"):
        raise ValueError(
            "구형 .xls 는 읽을 수 없습니다. 엑셀에서 '다른 이름으로 저장' → "
            "'Excel 통합 문서(*.xlsx)' 로 바꿔서 올려주세요."
        )
    if lower.endswith(".csv") or lower.endswith(".txt"):
        return _read_csv(content)
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return _read_xlsx(content)
    # 확장자가 없거나 낯설면 내용으로 판단한다. xlsx 는 zip 이라 'PK' 로 시작한다.
    if content[:2] == b"PK":
        return _read_xlsx(content)
    return _read_csv(content)


def _read_csv(content: bytes) -> List[list]:
    text = None
    # 한국 회계 프로그램은 CSV 를 cp949 로 뱉는 일이 흔하다.
    for encoding in ("utf-8-sig", "cp949", "utf-8"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("파일 인코딩을 알 수 없습니다. UTF-8 또는 xlsx 로 저장해 주세요.")
    sample = text[:4096]
    delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [list(row) for row in reader]


def _read_xlsx(content: bytes) -> List[list]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - 배포 이미지에는 항상 있다
        raise ValueError("서버에 엑셀 리더가 설치되어 있지 않습니다(openpyxl).")
    try:
        # data_only=True 라야 수식 대신 계산된 값을 읽는다.
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise ValueError("엑셀 파일을 열 수 없습니다. 손상되었거나 지원하지 않는 형식입니다.")
    try:
        ws = wb.worksheets[0]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _col_label(index: int) -> str:
    label = ""
    n = index + 1
    while n > 0:
        n, rem = divmod(n - 1, 26)
        label = chr(65 + rem) + label
    return label


def _find_article10_col(cells, name_col, amount_col) -> Optional[int]:
    """머리글 행에서 ⑩ 과세제외액 열을 찾는다. 없으면 None."""
    for i, cell in enumerate(cells):
        if i in (name_col, amount_col) or cell is None:
            continue
        text = str(cell).strip().lower()
        if text and any(h in text for h in ARTICLE10_HEADERS):
            return i
    return None


def extract_entries(rows: List[list]) -> Tuple[List[Tuple[str, int, int]], List[str]]:
    """행 목록에서 (거래처명, 매출액, ⑩ 과세제외액) 을 뽑는다. 경고 문구도 함께 돌려준다.

    ⑩ 열은 선택이다. 머리글로 찾지 못하면 모두 0 으로 둔다 — ERP 에서 그대로 뽑은
    파일에는 없는 것이 정상이고, 없다고 해서 업로드를 막을 이유가 없다.
    """
    warnings: List[str] = []
    if not rows:
        raise ValueError("파일에 읽을 내용이 없습니다.")

    name_col = amount_col = article10_col = None
    start_row = 0
    for i, cells in enumerate(rows[:HEADER_SCAN_ROWS]):
        score, n_col, a_col = _header_score(cells)
        if score == 2:
            name_col, amount_col, start_row = n_col, a_col, i + 1
            article10_col = _find_article10_col(cells, n_col, a_col)
            break

    if name_col is None or amount_col is None:
        name_col, amount_col = _sniff_columns(rows)
        start_row = 0
        if name_col is None or amount_col is None:
            raise ValueError(
                "거래처명 열과 매출액 열을 찾지 못했습니다. "
                "'양식 다운로드' 로 받은 파일에 금액만 채워서 올리면 확실합니다."
            )
        warnings.append(
            f"머리글을 찾지 못해 {_col_label(name_col)}열을 거래처명, "
            f"{_col_label(amount_col)}열을 매출액으로 읽었습니다. 결과를 확인해 주세요."
        )

    entries: List[Tuple[str, int, int]] = []
    skipped_no_amount = 0
    data_rows = rows[start_row:]
    if len(data_rows) > MAX_DATA_ROWS:
        warnings.append(f"행이 너무 많아 앞의 {MAX_DATA_ROWS}행만 읽었습니다.")
        data_rows = data_rows[:MAX_DATA_ROWS]

    for cells in data_rows:
        if not cells or all(c is None or str(c).strip() == "" for c in cells):
            continue
        if _is_total_row(cells):
            continue
        raw_name = cells[name_col] if name_col < len(cells) else None
        raw_amount = cells[amount_col] if amount_col < len(cells) else None
        name = str(raw_name).strip() if raw_name is not None else ""
        if not name or not normalize_name(name):
            continue
        amount = parse_amount(raw_amount)
        if amount is None:
            skipped_no_amount += 1
            continue
        article10 = 0
        if article10_col is not None and article10_col < len(cells):
            article10 = parse_amount(cells[article10_col]) or 0
        entries.append((name, amount, max(article10, 0)))

    if not entries:
        raise ValueError("거래처와 금액이 함께 있는 행을 찾지 못했습니다. 파일 내용을 확인해 주세요.")
    if skipped_no_amount:
        warnings.append(f"금액을 숫자로 읽을 수 없는 {skipped_no_amount}개 행은 건너뛰었습니다.")
    return entries, warnings


def match_entries(entries: List[Tuple[str, int, int]], companies: List[str]) -> dict:
    """뽑아낸 (이름, 금액) 을 서버 법인 목록에 맞춘다.

    같은 법인이 여러 줄에 나오면 합산한다(지점·월별로 쪼개진 파일이 흔하다).
    맞추지 못한 것은 버리지 않고 그대로 돌려준다 — 호출부가 사용자에게 보여준다.
    """
    lookup: Dict[str, str] = {}
    for company in companies:
        key = normalize_name(company)
        if key:
            lookup.setdefault(key, company)

    # 별칭은 실제 법인명을 덮지 않는다(setdefault). 가리키는 법인이 목록에 없으면 건너뛴다.
    for alias, canonical in NAME_ALIASES.items():
        target = lookup.get(normalize_name(canonical))
        if target:
            lookup.setdefault(normalize_name(alias), target)

    matched: Dict[str, int] = {}
    matched_article10: Dict[str, int] = {}
    matched_sources: Dict[str, List[str]] = {}
    unmatched: Dict[str, int] = {}
    unmatched_display: Dict[str, str] = {}
    negatives: List[str] = []

    for name, amount, article10 in entries:
        if amount < 0:
            negatives.append(name)
        key = normalize_name(name)
        company = lookup.get(key)
        if company:
            matched[company] = matched.get(company, 0) + amount
            matched_article10[company] = matched_article10.get(company, 0) + article10
            sources = matched_sources.setdefault(company, [])
            if name not in sources:
                sources.append(name)
        else:
            unmatched[key] = unmatched.get(key, 0) + amount
            unmatched_display.setdefault(key, name)

    warnings: List[str] = []
    if negatives:
        shown = ", ".join(negatives[:3])
        more = f" 외 {len(negatives) - 3}건" if len(negatives) > 3 else ""
        warnings.append(f"음수 금액이 있습니다: {shown}{more}. 반품·취소분인지 확인하세요.")

    return {
        "matched": [
            {"company": c, "amount": matched[c], "sources": matched_sources[c],
             "article10": matched_article10.get(c, 0)}
            for c in sorted(matched, key=lambda x: -matched[x])
        ],
        "unmatched": [
            {"name": unmatched_display[k], "amount": unmatched[k]}
            for k in sorted(unmatched, key=lambda x: -unmatched[x])
        ],
        "warnings": warnings,
    }


def import_related_sales(content: bytes, filename: str, companies: List[str]) -> dict:
    """업로드 파일 → 화면이 그대로 쓸 수 있는 매칭 결과."""
    rows = read_rows(content, filename)
    entries, read_warnings = extract_entries(rows)
    result = match_entries(entries, companies)
    result["warnings"] = read_warnings + result["warnings"]
    result["stats"] = {
        "rows_read": len(entries),
        "matched_count": len(result["matched"]),
        "unmatched_count": len(result["unmatched"]),
        "matched_total": sum(m["amount"] for m in result["matched"]),
        "unmatched_total": sum(u["amount"] for u in result["unmatched"]),
        "article10_total": sum(m["article10"] for m in result["matched"]),
    }
    return result


def build_template(companies: List[str]) -> bytes:
    """거래처명이 채워진 빈 양식. 사용자는 금액과 (해당되면) ⑩ 과세제외액을 채워 올린다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "특수관계자 세부매출"

    ws.append(["거래처명", "매출액(원)", "제10항 과세제외액(원)"])
    header_fill = PatternFill("solid", fgColor="EAF1FA")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for company in companies:
        ws.append([company, None])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = "#,##0"

    # 거래처명을 고쳐 쓰면 매칭이 깨지므로 파일 안에도 한 번 더 적어둔다.
    note_row = ws.max_row + 2
    note = ws.cell(row=note_row, column=1,
                   value="※ B열 금액만 채워서 그대로 올려주세요. 거래처명(A열)은 수정하지 마세요.")
    note.font = Font(color="64748B", size=10)
    note2 = ws.cell(row=note_row + 1, column=1,
                    value="※ C열(제10항 과세제외액)은 해당하는 거래처만 채우고, 없으면 비워 두세요.")
    note2.font = Font(color="64748B", size=10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
