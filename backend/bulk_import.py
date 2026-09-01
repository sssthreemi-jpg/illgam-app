"""통합본(법인 1개 = 시트 1개) 엑셀을 읽어 법인별 판정 입력으로 바꾼다.

단일 법인 업로드(`excel_import`)는 '거래처 / 금액' 표 하나만 읽으면 됐지만,
통합본은 시트마다 법인 기본정보(총매출·영업이익)와 매출 상세표가 같이 들어 있고
**시트마다 표가 시작하는 열이 다르다**(대웅제약은 D/G, 대웅개발은 C/F).
그래서 열 번호를 고정하지 않고 라벨을 찾아 그 오른쪽/아래를 읽는다.

**연환산 열이 이 파일의 정본이다.** 원본 셀 수식이 `=(E15-F15)*$M$5/$N$5` 처럼
반기 금액에 계수를 곱하고, 총매출·영업이익도 같은 계수를 쓴다. 그래서 거래처 매출은
그 열(`매출 합계 (1년 환산)`)을 그대로 읽고, 총매출·영업이익도 같은 열에서 읽는다 —
분자와 분모의 기준이 파일 안에서 이미 맞아 있다. 이 열에는 해외매출이 이미 빠져 있어
⑩ 으로 다시 넘기지 않는다(두 번 빼면 판정비율이 실제보다 낮아진다).

라벨 오른쪽 '첫' 칸을 값으로 삼으면 안 된다. 라벨과 연환산 값 사이에 안내 문구와
반기 금액이 끼어 있어(`총매출액 | '반기매출입력' | 454,217,451 | 908,434,902`),
첫 칸만 보면 문구를 값으로 읽고 데이터가 있는 법인을 통째로 입력대기로 빼버린다.

연환산 열이 없는 파일을 위해 계수를 역산하는(환산합계 ÷ (실매출 − 해외매출))
대체 경로를 남겨 뒀다. 그 경로에서는 해외매출을 ⑩ 으로 넘긴다.

이 모듈은 **아무것도 저장하지 않고 계산도 하지 않는다.** 판정은 calc 가 한다.
"""

import io
from typing import Dict, List, Optional

from backend.excel_import import (
    MAX_UPLOAD_BYTES,
    NAME_ALIASES,
    _is_total_row,
    normalize_name,
    parse_amount,
)

# 시트 1개가 법인 1개다. 통합본이 이보다 커지면 사람이 만든 파일이 아닐 가능성이 높다.
MAX_SHEETS = 200
# 기본정보 라벨은 표 위쪽에만 있다. 이 범위를 벗어나면 찾지 않는다.
HEADER_SCAN_ROWS = 14
HEADER_SCAN_COLS = 20
# 라벨 오른쪽으로 값을 찾을 때 몇 칸까지 볼지(병합셀 때문에 바로 옆이 아닐 수 있다).
VALUE_SCAN_COLS = 7
# 상세표에서 읽을 최대 행 수.
MAX_DETAIL_ROWS = 2000

# 기본정보 라벨. 정확히 일치하는 셀을 먼저 찾고, 없으면 포함으로 한 번 더 본다.
FIELD_LABELS = {
    "company_name": ("법인명", "회사명", "법인"),
    "size": ("기업구분", "기업규모", "규모"),
    "total_sales": ("총매출액", "총매출", "매출액계", "매출총액"),
    "operating_income": ("영업이익",),
    # 통합본 원본에는 없다. 나중에 양식에 추가되면 이 라벨로 자동으로 읽힌다.
    "corporate_tax": ("법인세상당액", "법인세"),
}

# 상세표 헤더. '해외매출' 을 '매출액' 보다 먼저 봐야 한다 —
# 실제 헤더가 '매출액 중 해외매출(L/C, 내국신용장 등)' 이라 '매출액' 으로도 걸린다.
DETAIL_NAME_HEADERS = ("매출거래처", "거래처")
DETAIL_FOREIGN_HEADERS = ("해외매출",)
DETAIL_TOTAL_HEADERS = ("1년환산", "매출합계", "합계")
DETAIL_AMOUNT_HEADERS = ("매출액", "매출")

# 상세표 헤더 아래에 붙는 설명 행. 거래처명이 아니다.
LEGEND_TEXTS = {"a", "b", "(a-b)", "발생한실제매출", "실제매출", "금액"}

# 표 아래에 붙는 집계 행. 이름 칸에 '특관매출 비율' 같은 문구가 들어가고 값이 0.9987 처럼
# 비율이라, 그대로 두면 금액 1원짜리 거래처로 잡힌다.
SUMMARY_MARKERS = ("비율", "합계", "소계", "총계", "검토", "판정")


def _text(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def _squash(value) -> str:
    """라벨 비교용. 공백·줄바꿈을 없애고 소문자로."""
    return "".join(_text(value).split()).lower()


def _find_label(rows, labels) -> Optional[tuple]:
    """(행 index, 열 index). 정확히 일치하는 셀을 우선한다."""
    squashed = [_squash(l) for l in labels]
    for exact in (True, False):
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                if not isinstance(cell, str):
                    continue
                key = _squash(cell)
                if not key:
                    continue
                for label in squashed:
                    if (key == label) if exact else (label in key):
                        return i, j
    return None


def _value_right_of(rows, pos, value_col=None):
    """라벨 셀의 값. 연환산 열(`매출 합계 (1년 환산)`)을 정본으로 본다.

    라벨 오른쪽 '첫' 칸을 값으로 삼으면 안 된다. 실제 파일에는 라벨과 값 사이에
    안내 문구와 반기 금액이 끼어 있다:

        C6=총매출액 | D6='반기매출입력' | E6=454,217,451(반기) | F6=908,434,902(연환산)

    첫 칸만 보면 '반기매출입력'을 값으로 읽어 **데이터가 있는 법인을 입력대기로
    빼버린다.** 실제로 아이엔·대웅테라가 그렇게 빠졌다.

    열을 F 로 고정할 수도 없다 — 대웅·대웅제약은 표가 한 칸 오른쪽에서 시작해
    같은 값이 G 에 있고 F 는 비어 있다. 그래서 그 시트의 상세표 헤더에서 찾은
    연환산 열(`value_col`)을 쓰고, 없으면 오른쪽에서 가장 먼 숫자를 고른다.
    """
    if pos is None:
        return None
    i, j = pos
    row = rows[i]
    # 1순위: 그 시트의 연환산 열. 총매출·영업이익도 상세표와 같은 열에 놓여 있다.
    if value_col is not None and j < value_col < len(row):
        if parse_amount(row[value_col]) is not None:
            return row[value_col]
    # 2순위: 라벨 오른쪽에서 가장 먼 숫자(반기 금액 왼쪽에 연환산이 오는 일은 없다).
    found = None
    for k in range(j + 1, min(j + 1 + VALUE_SCAN_COLS, len(row))):
        if parse_amount(row[k]) is not None:
            found = row[k]
    if found is not None:
        return found
    # 숫자가 하나도 없으면 안내 문구라도 돌려준다 — 왜 못 읽었는지 화면에 적어야 한다.
    for k in range(j + 1, min(j + 1 + VALUE_SCAN_COLS, len(row))):
        if row[k] not in (None, ""):
            return row[k]
    return None


def _find_detail_header(rows) -> Optional[dict]:
    """상세표 헤더 행을 찾아 열 위치를 돌려준다."""
    for i, row in enumerate(rows):
        name_col = None
        for j, cell in enumerate(row):
            if isinstance(cell, str) and any(h in _squash(cell) for h in DETAIL_NAME_HEADERS):
                name_col = j
                break
        if name_col is None:
            continue
        cols = {"row": i, "name": name_col, "amount": None, "foreign": None, "total": None}
        for j in range(name_col + 1, len(row)):
            cell = row[j]
            if not isinstance(cell, str):
                continue
            key = _squash(cell)
            if not key:
                continue
            # 순서가 중요하다. 해외 → 환산합계 → 매출액.
            if cols["foreign"] is None and any(h in key for h in DETAIL_FOREIGN_HEADERS):
                cols["foreign"] = j
            elif cols["total"] is None and any(h in key for h in DETAIL_TOTAL_HEADERS):
                cols["total"] = j
            elif cols["amount"] is None and any(h in key for h in DETAIL_AMOUNT_HEADERS):
                cols["amount"] = j
        if cols["amount"] is not None or cols["total"] is not None:
            return cols
    return None


def _annualize_factor(entries) -> float:
    """환산합계 ÷ (실매출 − 해외매출). 파일 전체가 같은 계수를 쓰므로 최빈값을 고른다."""
    counts: Dict[float, int] = {}
    for e in entries:
        base = (e["amount"] or 0) - (e["foreign"] or 0)
        total = e["total"]
        if not base or total is None:
            continue
        factor = round(total / base, 4)
        if factor <= 0:
            continue
        counts[factor] = counts.get(factor, 0) + 1
    if not counts:
        return 1.0
    return max(counts.items(), key=lambda kv: kv[1])[0]


def _read_detail(rows, cols) -> List[dict]:
    out = []
    for row in rows[cols["row"] + 1:cols["row"] + 1 + MAX_DETAIL_ROWS]:
        if cols["name"] >= len(row):
            continue
        name = _text(row[cols["name"]])
        key = _squash(name)
        if len(name) < 2 or key in LEGEND_TEXTS:
            continue
        if any(m in key for m in SUMMARY_MARKERS):
            continue
        if _is_total_row(row):
            continue

        def at(key):
            col = cols.get(key)
            if col is None or col >= len(row):
                return None
            return parse_amount(row[col])

        amount, foreign, total = at("amount"), at("foreign"), at("total")
        if amount is None and total is None:
            continue
        out.append({"name": name, "amount": amount, "foreign": foreign or 0, "total": total})
    return out


def _lookup_table(companies: List[str]) -> Dict[str, str]:
    lookup: Dict[str, str] = {}
    for company in companies:
        key = normalize_name(company)
        if key:
            lookup.setdefault(key, company)
    for alias, canonical in NAME_ALIASES.items():
        target = lookup.get(normalize_name(canonical))
        if target:
            lookup.setdefault(normalize_name(alias), target)
    return lookup


def _parse_sheet(title, rows, lookup, sizes) -> dict:
    warnings: List[str] = []
    head = [list(r[:HEADER_SCAN_COLS]) for r in rows[:HEADER_SCAN_ROWS]]

    # 기본정보 값은 상세표의 연환산 열에 놓이므로 헤더를 먼저 찾는다.
    cols = _find_detail_header(rows)
    value_col = cols["total"] if cols else None
    if cols is None:
        warnings.append("특수관계자 매출 상세표를 찾지 못했습니다. '매출거래처' 머리글이 있는지 확인하세요.")

    raw_name = _value_right_of(head, _find_label(head, FIELD_LABELS["company_name"]))
    # 법인명 셀이 정본이고, 없으면 시트명으로 맞춰본다(시트명은 '생명과학' 처럼 줄임말이 많다).
    company = lookup.get(normalize_name(raw_name)) or lookup.get(normalize_name(title))
    size_excel = _text(_value_right_of(head, _find_label(head, FIELD_LABELS["size"])))

    def money(field):
        raw = _value_right_of(head, _find_label(head, FIELD_LABELS[field]), value_col)
        return parse_amount(raw), raw

    total_sales, total_raw = money("total_sales")
    operating_income, income_raw = money("operating_income")
    corporate_tax, _ = money("corporate_tax")

    entries = _read_detail(rows, cols) if cols else []
    factor = _annualize_factor(entries)
    related: Dict[str, int] = {}
    article10: Dict[str, int] = {}
    unmatched: List[dict] = []
    foreign_total = 0
    for e in entries:
        if e["total"] is not None:
            # 연환산 열(=(A−B)×계수)이 이 파일의 정본이다. 해외매출이 이미 빠져 있고
            # 연환산도 끝나 있어, 우리가 계수를 역산해 곱하는 것보다 정확하다.
            amount, foreign = e["total"], 0
        else:
            # 연환산 열이 없는 파일용 대체 경로. 실매출에 계수를 곱하고
            # 해외매출은 ⑩ 으로 넘긴다.
            amount = round((e["amount"] or 0) * factor)
            foreign = round(e["foreign"] * factor)
        foreign_total += round(e["foreign"] * factor)
        if amount <= 0 and foreign <= 0:
            continue
        target = lookup.get(normalize_name(e["name"]))
        if not target:
            unmatched.append({"name": e["name"], "amount": amount})
            continue
        related[target] = related.get(target, 0) + amount
        if foreign:
            article10[target] = article10.get(target, 0) + foreign

    # 규모는 앱 데이터가 정본이다. 엑셀 표기('일반기업','대기업','중소기업')는 참고만 한다.
    size_app = sizes.get(company) if company else None
    # 기업구분 칸이 비었거나 0 인 시트가 있다. 그건 '다르다'가 아니라 '안 적혔다' 이므로
    # 경고를 띄우면 진짜 불일치(대웅바이오: 엑셀 대기업 vs 서버 일반)가 묻힌다.
    size_written = bool(size_excel) and parse_amount(size_excel) is None
    size_mismatch = bool(
        company and size_written and size_app
        and not normalize_name(size_excel).startswith(normalize_name(size_app))
    )

    status = "ok"
    if company is None and cols is None:
        # 매출 상세표도 없고 법인도 못 맞춘 시트. 통합본에는 '특관매출 요약', '기업분류',
        # '0. 보고' 처럼 법인 시트가 아닌 것이 섞여 있다. 이런 시트까지 '미매칭'으로
        # 세면 진짜 확인이 필요한 법인이 그 안에 묻힌다.
        status = "건너뜀"
        warnings = ["법인 시트가 아닙니다(매출 상세표가 없습니다). 판정에서 제외합니다."]
    elif company is None:
        status = "미매칭"
        warnings.append(
            "서버 법인 목록에서 찾지 못한 이름입니다: " + (_text(raw_name) or title)
        )
    elif total_sales is None or operating_income is None:
        status = "입력대기"
        missing = []
        if total_sales is None:
            missing.append("총매출(" + (_text(total_raw) or "빈칸") + ")")
        if operating_income is None:
            missing.append("영업이익(" + (_text(income_raw) or "빈칸") + ")")
        warnings.append("아직 채워지지 않았습니다: " + ", ".join(missing))
    else:
        related_total = sum(related.values())
        if total_sales <= 0:
            status = "입력대기"
            warnings.append("총매출이 0 이하라 비율을 계산할 수 없습니다.")
        elif related_total > total_sales:
            status = "확인필요"
            warnings.append(
                "특수관계자 매출({:,}원)이 총매출({:,}원)보다 큽니다. "
                "총매출이 아직 임시값인지 확인하세요.".format(related_total, total_sales)
            )

    if size_mismatch:
        warnings.append(
            "기업구분이 다릅니다 — 엑셀 '{}', 서버 '{}'. "
            "정상거래비율이 달라지므로 서버 값으로 계산합니다.".format(size_excel, size_app)
        )
    if corporate_tax is None and status == "ok":
        warnings.append("법인세 상당액이 파일에 없습니다. 화면에서 입력하세요(미입력 시 세액이 과대 계산됩니다).")
    # 연환산 열을 쓰면 해외매출이 이미 빠져 있어 ⑩ 으로 따로 넘기지 않는다.
    # 금액이 묻히지 않게 얼마가 빠진 값인지는 남긴다.
    if foreign_total and not article10:
        warnings.append(
            "해외매출(L/C·내국신용장) {:,}원이 이미 빠진 '매출 합계(1년 환산)' 값입니다. "
            "제10항으로 따로 넘기지 않습니다.".format(foreign_total))

    return {
        "sheet": title,
        "company": company,
        "excel_name": _text(raw_name) or title,
        "status": status,
        "size_excel": size_excel,
        "size_app": size_app,
        "size_mismatch": size_mismatch,
        "total_sales": total_sales,
        "operating_income": operating_income,
        "corporate_tax": corporate_tax,
        "annualize_factor": factor,
        "related_sales": related,
        "article10_exclusions": article10,
        "related_total": sum(related.values()),
        "article10_total": sum(article10.values()),
        # 연환산 열에서 이미 차감된 해외매출. 계산에는 안 쓰고 화면에만 보여준다.
        "foreign_sales_total": foreign_total,
        "counterparty_count": len(related),
        "unmatched": unmatched,
        "warnings": warnings,
    }


def parse_workbook(content: bytes, filename: str, companies: List[str],
                   sizes: Dict[str, str]) -> dict:
    """통합본 → 법인별 판정 입력. 계산도 저장도 하지 않는다."""
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError(
            "파일이 너무 큽니다. {}MB 이하로 올려주세요.".format(MAX_UPLOAD_BYTES // (1024 * 1024))
        )
    if not content:
        raise ValueError("빈 파일입니다.")
    lower = (filename or "").lower()
    if lower.endswith(".xls"):
        raise ValueError(
            "구형 .xls 는 읽을 수 없습니다. 엑셀에서 '다른 이름으로 저장' → "
            "'Excel 통합 문서(*.xlsx)' 로 바꿔서 올려주세요."
        )
    if not (lower.endswith(".xlsx") or lower.endswith(".xlsm") or content[:2] == b"PK"):
        raise ValueError("통합본은 시트가 여러 개인 엑셀(.xlsx)이어야 합니다.")

    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - 배포 이미지에는 항상 있다
        raise ValueError("서버에 엑셀 리더가 설치되어 있지 않습니다(openpyxl).")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise ValueError("엑셀 파일을 열 수 없습니다. 손상되었거나 지원하지 않는 형식입니다.")

    lookup = _lookup_table(companies)
    sheets = []
    warnings: List[str] = []
    try:
        titles = wb.sheetnames
        if len(titles) > MAX_SHEETS:
            warnings.append(
                "시트가 {}개를 넘어 앞의 {}개만 읽었습니다.".format(MAX_SHEETS, MAX_SHEETS))
            titles = titles[:MAX_SHEETS]
        for title in titles:
            ws = wb[title]
            rows = [list(r) for r in ws.iter_rows(
                max_row=HEADER_SCAN_ROWS + MAX_DETAIL_ROWS, values_only=True)]
            if not any(any(c not in (None, "") for c in r) for r in rows):
                continue
            sheets.append(_parse_sheet(title, rows, lookup, sizes))
    finally:
        wb.close()

    if not sheets:
        raise ValueError("읽을 수 있는 시트가 없습니다.")

    seen: Dict[str, List[str]] = {}
    for s in sheets:
        if s["company"]:
            seen.setdefault(s["company"], []).append(s["sheet"])
    for company, titles in seen.items():
        if len(titles) > 1:
            warnings.append(
                "{} 시트가 {}개입니다({}). 모두 계산합니다.".format(
                    company, len(titles), ", ".join(titles)))

    # 판정 대상 법인 중 시트가 아예 없는 곳. 미제출을 놓치지 않도록 알린다.
    missing = [c for c in companies if c in sizes and c not in seen]

    return {
        "sheets": sheets,
        "missing_companies": missing,
        "warnings": warnings,
        "stats": {
            "sheets_read": len(sheets),
            "ready": sum(1 for s in sheets if s["status"] == "ok"),
            "pending": sum(1 for s in sheets if s["status"] not in ("ok", "건너뜀")),
            "skipped": sum(1 for s in sheets if s["status"] == "건너뜀"),
            "missing": len(missing),
        },
    }
