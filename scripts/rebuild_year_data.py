# -*- coding: utf-8 -*-
"""연도별 지분 데이터(JSON)를 원본 지분율 엑셀에서 다시 만든다.

`backend/data/<연도>/` 는 기밀이라 git 에 없다. 이 스크립트가 있으면 엑셀만 있어도
`shareholder_holdings.json` 과 `intercompany_holdings.json` 두 개는 복구된다.

**나머지 세 파일은 이 스크립트로 복구되지 않는다.**
  - `company_sizes.json`  : 기업분류 엑셀이 연도별로 따로 있고, 손으로 고친 값이 섞여 있다.
                            여기서는 **법인 목록의 정본으로 읽기만** 한다.
  - `params.json`         : 세율·비율 등 직접 관리하는 값.
  - `section18_indirect_investors.json` : 세무 검토로 확정해 등재하는 값.
  이 세 개는 별도로 백업해 둘 것.

재생성 전에 `3.지배주주지분율 요약` 이 `1.직접지분율` 과 맞는지 검증한다. 요약만 낡은
파일이 실제로 있었고(26.06말 260730), 그대로 쓰면 시지바이오 지분이 0% 가 된다.

사용법 (기본은 미리보기, 실제로 덮어쓰려면 --write):

    python scripts/rebuild_year_data.py --year 2025 \
        --holdings "C:/.../25.12말 기준 일감 증여세 (지분율)_260713.xlsx"

    python scripts/rebuild_year_data.py --year 2025 --holdings "..." --write
"""

import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 이 필요합니다:  pip install openpyxl")

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

SHEET_SUMMARY = "3.지배주주지분율 요약"
SHEET_DIRECT = "1.직접지분율"

# 지배주주 코드와 `3.지배주주지분율 요약` 시트의 '합계' 열(0-based).
# 시트는 코드마다 직접/간접/합계 3열이 반복된다: F=A합계, I=B합계, L=C합계 …
CODES = ["A", "B", "C", "D", "C1", "C11", "C12"]
SUM_COL = {"A": 5, "B": 8, "C": 11, "D": 14, "C1": 17, "C11": 20, "C12": 23}
NAME_COL = 1  # B열 법인명

# `1.직접지분율` 시트: A열 출자회사, B열 피출자회사, F열 지분율
OWNER_COL, TARGET_COL, RATIO_COL = 0, 1, 5

# 지분율 엑셀은 4개 시트 모두 `엠베이스`, 기업분류 엑셀과 앱 데이터는 `시지엠베이스`.
# 기업분류 쪽 이름이 정본이므로 그쪽으로 맞춘다.
ALIAS = {"엠베이스": "시지엠베이스"}

ROUND = 10  # 기존 JSON 과 같은 자릿수

# --- `3.지배주주지분율 요약` 검증 문턱 ------------------------------------------
# 요약 시트는 `1.직접지분율` 을 체인으로 전개한 결과여야 한다. 그런데 요약만 낡은 채로
# 오는 파일이 있다 — `26.06말 ..._260730.xlsx` 는 제목이 2025.12.31 이고 시지 계열 등
# 10개 법인의 합계가 0% 다. 그대로 재생성하면 시지바이오 지분이 80.5% → 0% 가 되어
# 증여세가 조용히 0 이 된다. 그래서 직접 계산해 대조하고, 심하면 멈춘다.
MATCH_TOL = 0.001             # 0.1%p 이내면 같은 값으로 본다(반올림·자기주식 처리 차이).
ZERO_FATAL = 0.01             # 요약은 0 인데 체인은 1%p 이상 — 계산이 누락된 것이다.
MISMATCH_FATAL_RATIO = 0.25   # 요약의 1/4 넘게 어긋나면 시트 자체를 믿을 수 없다.
CHAIN_ROUNDS = 200            # 체인 전개 반복 상한(보통 5~6회면 수렴한다).


def norm(v):
    if not isinstance(v, str):
        return None
    v = v.strip()
    return ALIAS.get(v, v) or None


def read_shareholders(ws, allowed=None):
    """{법인: {A..C12, sum}} — 코드별 '합계' 열을 그대로 읽는다.

    allowed 를 주면 그 법인만 남긴다. None 이면 시트에 있는 대로 전부 읽는다
    (요약 시트 검증은 법인 목록 밖 법인까지 봐야 해서 필터 없이 읽는다).
    """
    out = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = norm(row[NAME_COL] if len(row) > NAME_COL else None)
        if not name or (allowed is not None and name not in allowed):
            continue
        vals = {}
        for code in CODES:
            col = SUM_COL[code]
            v = row[col] if len(row) > col else None
            if not isinstance(v, (int, float)):
                vals = None
                break
            vals[code] = float(v)
        if vals is None:
            continue
        rounded = {c: round(vals[c], ROUND) for c in CODES}
        # 합계는 반올림한 코드별 값을 더한다(원본 JSON 과 같은 순서로 계산해야 끝자리가 맞는다).
        rounded["sum"] = round(sum(rounded[c] for c in CODES), ROUND)
        out[name] = rounded
    return out


def read_direct(ws):
    """`1.직접지분율` 원시 행 (출자, 피출자, 지분율). 필터를 걸지 않는다.

    체인 계산에는 개인 코드(A·C11 …) 행과 법인 목록 밖 중간 지주사도 필요해서,
    거르는 일은 이 함수를 쓰는 쪽에서 한다.
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        owner = norm(row[OWNER_COL] if len(row) > OWNER_COL else None)
        target = norm(row[TARGET_COL] if len(row) > TARGET_COL else None)
        ratio = row[RATIO_COL] if len(row) > RATIO_COL else None
        if not owner or not target or not isinstance(ratio, (int, float)):
            continue
        rows.append((owner, target, float(ratio)))
    return rows


def read_intercompany(direct_rows, order):
    """{출자법인: {피출자법인: 지분율}} — 개인 코드 행은 법인 목록 필터에서 걸러진다."""
    allowed = set(order)
    out = {c: {} for c in order}
    for owner, target, ratio in direct_rows:
        if owner not in allowed or target not in allowed:
            continue
        out[owner][target] = round(ratio, ROUND)
    return out


def direct_graph(direct_rows):
    """{출자자: {피출자: 지분율}} — 개인 코드까지 포함한 전체 그래프."""
    edges = {}
    for owner, target, ratio in direct_rows:
        edges.setdefault(owner, {})
        edges[owner][target] = edges[owner].get(target, 0.0) + ratio
    return edges


def chain_holdings(edges, codes=CODES):
    """지배주주별 직·간접 지분율을 직접지분율 그래프에서 전개한다.

    hold[p][c] = 직접(p→c) + Σ_o 지분율(o→c) × hold[p][o]
    순환출자가 있어도 값이 수렴하도록 반복해서 푼다.
    """
    hold = {p: {} for p in codes}
    for _ in range(CHAIN_ROUNDS):
        moved = 0.0
        for p in codes:
            new = dict(edges.get(p, {}))
            for owner, targets in edges.items():
                if owner in codes:
                    continue
                share = hold[p].get(owner, 0.0)
                if not share:
                    continue
                for target, ratio in targets.items():
                    new[target] = new.get(target, 0.0) + ratio * share
            for k in set(new) | set(hold[p]):
                moved = max(moved, abs(new.get(k, 0.0) - hold[p].get(k, 0.0)))
            hold[p] = new
        if moved < 1e-15:
            break
    return hold


def summary_vs_direct(summary, hold):
    """요약 시트와 체인계산이 어긋나는 법인 목록 [(법인, 요약합계, 체인합계, 최대차)]."""
    out = []
    for co in sorted(summary):
        vals = summary[co]
        calc = {c: hold[c].get(co, 0.0) for c in CODES}
        worst = max(abs(calc[c] - vals[c]) for c in CODES)
        if worst > MATCH_TOL:
            out.append((co, sum(vals[c] for c in CODES), sum(calc.values()), worst))
    return out


def report_summary_check(summary, mismatches):
    """검증 결과를 찍고, 이대로 쓰면 안 되는 파일인지(True) 알려준다."""
    if not summary:
        print(f"[검증] `{SHEET_SUMMARY}` 에서 읽은 법인이 없습니다.")
        return True
    if not mismatches:
        print(f"요약 시트 검증: {len(summary)}개 법인 모두 "
              f"`{SHEET_DIRECT}` 체인계산과 일치합니다.")
        return False

    zeroed = [m for m in mismatches if abs(m[1]) < 1e-12 and m[2] >= ZERO_FATAL]
    ratio = len(mismatches) / len(summary)
    print(f"[검증] `{SHEET_SUMMARY}` 가 `{SHEET_DIRECT}` 체인계산과 어긋납니다 "
          f"— {len(summary)}개 중 {len(mismatches)}개 ({ratio:.0%})")
    for co, sheet_sum, calc_sum, _ in mismatches[:20]:
        print(f"      {co:<16} 요약 {sheet_sum * 100:7.3f}%"
              f"   직접지분율 체인 {calc_sum * 100:7.3f}%")
    if len(mismatches) > 20:
        print(f"      … 외 {len(mismatches) - 20}개")
    if zeroed:
        names = ", ".join(m[0] for m in zeroed)
        print(f"\n  → 요약은 0% 인데 실제로는 지분이 있는 법인 {len(zeroed)}개: {names}")
        print("     요약 시트의 계산이 누락된 것입니다. "
              "이대로 쓰면 이 법인들의 증여세가 조용히 0 이 됩니다.")
    return bool(zeroed) or ratio > MISMATCH_FATAL_RATIO


def dump(path, obj):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=1)
        print(file=f)  # 기존 파일처럼 끝에 개행 하나를 남긴다


def report(label, new, old):
    if old is None:
        print(f"  {label}: 기존 파일 없음 — 새로 만듭니다 ({len(new)}개 항목)")
        return
    if new == old:
        print(f"  {label}: 기존 파일과 동일 (변경 없음)")
        return
    added = sorted(set(new) - set(old))
    removed = sorted(set(old) - set(new))
    changed = [k for k in sorted(set(new) & set(old)) if new[k] != old[k]]
    print(f"  {label}: 추가 {len(added)} / 삭제 {len(removed)} / 변경 {len(changed)}")
    for k in added[:10]:
        print(f"      + {k}")
    for k in removed[:10]:
        print(f"      - {k}")
    for k in changed[:10]:
        print(f"      ~ {k}")
        print(f"          기존: {old[k]}")
        print(f"          신규: {new[k]}")


def main():
    p = argparse.ArgumentParser(description="지분율 엑셀에서 연도별 JSON 재생성")
    p.add_argument("--year", required=True, help="예: 2025")
    p.add_argument("--holdings", required=True, help="지분율 엑셀 경로")
    p.add_argument("--data-dir", default=None,
                   help="기본값: <repo>/backend/data/<연도>")
    p.add_argument("--write", action="store_true",
                   help="실제로 덮어쓴다. 없으면 미리보기만 한다.")
    p.add_argument("--skip-summary-check", action="store_true",
                   help="요약 시트 검증을 건너뛴다. 검증이 왜 걸렸는지 알고 있을 때만 쓸 것.")
    args = p.parse_args()

    out_dir = args.data_dir or os.path.join(REPO, "backend", "data", args.year)
    sizes_path = os.path.join(out_dir, "company_sizes.json")
    if not os.path.exists(sizes_path):
        sys.exit(
            f"{sizes_path} 가 없습니다.\n"
            "company_sizes.json 은 법인 목록의 정본이라 먼저 있어야 합니다 "
            "(이 스크립트로는 만들 수 없습니다 — 백업에서 복원하세요)."
        )

    with open(sizes_path, encoding="utf-8") as f:
        sizes = json.load(f)
    # `기타법인` 은 calc.company_list() 가 자동으로 붙이므로 데이터에 없어야 정상이다.
    # 순서까지 정본을 따라가야 재생성 결과가 기존 파일과 바이트 단위로 같아진다.
    order = [c for c in sizes if c != "기타법인"]
    allowed = set(order)
    print(f"법인 목록 정본: {sizes_path} ({len(allowed)}개)")

    wb = openpyxl.load_workbook(args.holdings, data_only=True, read_only=True)
    for sheet in (SHEET_SUMMARY, SHEET_DIRECT):
        if sheet not in wb.sheetnames:
            sys.exit(f"엑셀에 `{sheet}` 시트가 없습니다: {wb.sheetnames}")

    # 요약 시트는 두 번 읽지 않는다(read_only 워크시트라 재순회가 비싸다).
    # 검증에는 법인 목록 밖 법인도 필요해서 필터 없이 한 번 읽고 나중에 거른다.
    summary = read_shareholders(wb[SHEET_SUMMARY], allowed=None)
    direct_rows = read_direct(wb[SHEET_DIRECT])

    if args.skip_summary_check:
        print()
        print("[주의] --skip-summary-check: 요약 시트 검증을 건너뜁니다.")
    else:
        print()
        fatal = report_summary_check(
            summary, summary_vs_direct(summary, chain_holdings(direct_graph(direct_rows))))
        if fatal:
            print()
            print(f"`{SHEET_SUMMARY}` 시트를 그대로 쓸 수 없어 멈춥니다.")
            print(f"이 시트가 아니라 `{SHEET_DIRECT}` 이 정본입니다. "
                  "요약이 갱신된 파일을 받거나,")
            print("직접지분율 체인계산 값으로 대신할지 판단한 뒤 "
                  "--skip-summary-check 로 다시 실행하세요.")
            sys.exit(1)

    # 엑셀 행 순서가 아니라 정본 목록 순서로 맞춘다.
    shareholders = {c: summary[c] for c in order if c in summary}
    intercompany = read_intercompany(direct_rows, order)

    missing = sorted(allowed - set(shareholders))
    if missing:
        print(f"\n[경고] 기업분류에 있으나 지분율 엑셀에 없는 법인 {len(missing)}건: {missing}")

    print()
    targets = [
        ("shareholder_holdings.json", shareholders),
        ("intercompany_holdings.json", intercompany),
    ]
    for name, new in targets:
        path = os.path.join(out_dir, name)
        old = None
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                old = json.load(f)
        report(name, new, old)

    if not args.write:
        print("\n미리보기입니다. 실제로 쓰려면 --write 를 붙이세요.")
        return

    for name, new in targets:
        dump(os.path.join(out_dir, name), new)
    print(f"\n{out_dir} 에 2개 파일을 썼습니다.")
    print("params.json / section18_indirect_investors.json / company_sizes.json 은 건드리지 않았습니다.")


if __name__ == "__main__":
    main()
